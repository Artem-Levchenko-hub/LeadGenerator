"""Работа с лидами на Яндекс.Диске.

Архитектура хранения:
- `/Stenvik/leads.xlsx` — компактный дашборд: дата, компания, телефон, город,
  приоритет (с цветом), ссылка "Детали" (на MD-файл), + CRM-поля продажника
  (позвонил, статус сделки, ответственный, фидбэк).
- `/Stenvik/leads/<slug>.md` — подробности по каждой компании (описание, боли,
  услуги, хук для продажника) в красивом markdown-формате.

Каждый MD-файл публикуется отдельно → public URL вшивается в ячейку таблицы
как hyperlink, клик открывает файл в Яндекс.Документах.
"""
import logging
import re
import tempfile
import time
from datetime import datetime
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)

YANDEX_API = "https://cloud-api.yandex.net/v1/disk/resources"
LEADS_FOLDER = "/Stenvik/leads"

# Компактные колонки — всё что за пределами одного экрана ушло в MD
COLUMNS = [
    ("A", "Дата",             14),
    ("B", "Компания",          32),
    ("C", "Телефон",           18),
    ("D", "Город",             14),
    ("E", "Приоритет",         10),
    ("F", "Детали",            14),  # hyperlink на MD-файл
    # --- CRM-поля продажника ---
    ("G", "Позвонил",          12),  # dropdown ✓
    ("H", "Дата звонка",       14),
    ("I", "Ответственный",     16),
    ("J", "Статус сделки",     18),  # dropdown
    ("K", "Фидбэк по звонку",  46),
    # --- Технические ---
    ("L", "dedup_key",         20),  # скрытая
]

DEAL_STATUSES = [
    "Новый", "В работе", "Связались", "Квалифицирован",
    "Сделка", "Отказ", "Не наш",
]
CALL_VALUES = ["", "✓"]

PRIORITY_FILLS = {
    5: PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"),
    4: PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    3: PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid"),
    2: PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid"),
    1: PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="1A73E8", underline="single")
BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def _auth_headers() -> dict:
    if not settings.yandex_disk_token:
        raise RuntimeError("YANDEX_DISK_TOKEN не задан в .env")
    return {"Authorization": f"OAuth {settings.yandex_disk_token}"}


# ==== общие утилиты API ====

def _get_download_url(path: str) -> str | None:
    with httpx.Client(timeout=15.0) as client:
        r = client.get(f"{YANDEX_API}/download", params={"path": path}, headers=_auth_headers())
        if r.status_code == 404:
            return None
        r.raise_for_status()
        return r.json()["href"]


def _get_upload_url(path: str, overwrite: bool = True) -> str:
    with httpx.Client(timeout=15.0) as client:
        r = client.get(
            f"{YANDEX_API}/upload",
            params={"path": path, "overwrite": str(overwrite).lower()},
            headers=_auth_headers(),
        )
        r.raise_for_status()
        return r.json()["href"]


def _ensure_parent_dir(path: str) -> None:
    parts = [p for p in path.strip("/").split("/")[:-1] if p]
    if not parts:
        return
    current = ""
    with httpx.Client(timeout=15.0) as client:
        for p in parts:
            current = f"{current}/{p}" if current else f"/{p}"
            r = client.put(YANDEX_API, params={"path": current}, headers=_auth_headers())
            if r.status_code not in (201, 409, 423):
                r.raise_for_status()


def _upload_bytes(path: str, data: bytes, max_retries: int = 6) -> None:
    """Загружает байты на Я.Диск по пути. Ретраит 423 LOCKED с exp backoff."""
    _ensure_parent_dir(path)
    last_err = None
    for attempt in range(max_retries):
        try:
            upload_url = _get_upload_url(path, overwrite=True)
            with httpx.Client(timeout=120.0) as client:
                r = client.put(upload_url, content=data)
                r.raise_for_status()
            return
        except httpx.HTTPStatusError as e:
            last_err = e
            if e.response.status_code == 423:
                delay = min(2 ** attempt, 20)
                logger.warning("Upload locked (423) %s — retry %s/%s in %ss", path, attempt + 1, max_retries, delay)
                time.sleep(delay)
                continue
            raise
    raise RuntimeError(f"Upload {path} failed after {max_retries} retries: {last_err}")


def _publish_and_get_url(path: str) -> str:
    """Публикует файл/папку, возвращает public_url."""
    with httpx.Client(timeout=15.0) as client:
        r = client.put(f"{YANDEX_API}/publish", params={"path": path}, headers=_auth_headers())
        r.raise_for_status()
        r2 = client.get(
            YANDEX_API,
            params={"path": path, "fields": "public_url"},
            headers=_auth_headers(),
        )
        r2.raise_for_status()
        return r2.json().get("public_url", "")


# ==== MD файлы деталей лида ====

def _slugify(name: str) -> str:
    """Безопасный slug для имени файла из названия компании."""
    # транслит: базовый, достаточный для имени файла
    tr = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    s = name.lower().strip()
    s = "".join(tr.get(ch, ch) for ch in s)
    s = re.sub(r"[^a-z0-9\- ]+", "", s)
    s = re.sub(r"\s+", "-", s).strip("-")
    return s[:80] or "company"


def _build_markdown(
    *,
    company_name: str,
    website_url: str | None,
    city: str | None,
    industry: str | None,
    website_status: str,
    phone: str | None,
    priority: int,
    priority_reason: str,
    summary: str,
    pains: list[str],
    recommended_services: list[str],
    sales_hook: str,
) -> str:
    """Красиво форматирует данные лида в markdown."""
    priority_emoji = {5: "🔴", 4: "🟠", 3: "🟢", 2: "⚪", 1: "⚪"}.get(priority, "⚪")

    lines: list[str] = []
    lines.append(f"# {company_name}")
    lines.append("")
    lines.append(f"**Приоритет:** {priority_emoji} {priority}/5 — {priority_reason}")
    lines.append("")
    lines.append("## Реквизиты")
    lines.append("")
    if website_url:
        lines.append(f"- **Сайт:** [{website_url}]({website_url})")
    if phone:
        lines.append(f"- **Телефон:** {phone}")
    if city:
        lines.append(f"- **Город:** {city}")
    if industry:
        lines.append(f"- **Индустрия:** {industry}")
    lines.append(f"- **Статус сайта:** {website_status}")
    lines.append(f"- **Дата анализа:** {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("## О компании")
    lines.append("")
    lines.append(summary)
    lines.append("")
    lines.append("## Боли (что не так с точки зрения цифровизации)")
    lines.append("")
    for i, pain in enumerate(pains, 1):
        lines.append(f"{i}. {pain}")
    lines.append("")
    lines.append("## Что предложить (услуги Stenvik)")
    lines.append("")
    for s in recommended_services:
        lines.append(f"- {s}")
    lines.append("")
    lines.append("## Хук для продажника")
    lines.append("")
    lines.append(f"> {sales_hook}")
    lines.append("")
    return "\n".join(lines)


def upload_markdown_lead(company_name: str, markdown: str) -> tuple[str, str]:
    """Загружает MD-файл в /Stenvik/leads/ и публикует. Возвращает (remote_path, public_url)."""
    slug = _slugify(company_name)
    # чтобы избежать коллизий — добавим timestamp-суффикс
    stamp = datetime.now().strftime("%y%m%d%H%M")
    remote_path = f"{LEADS_FOLDER}/{slug}-{stamp}.md"
    _upload_bytes(remote_path, markdown.encode("utf-8"))
    public_url = _publish_and_get_url(remote_path)
    logger.info("Uploaded + published MD: %s → %s", remote_path, public_url)
    return remote_path, public_url


def upload_html_lead(company_name: str, html: str) -> tuple[str, str]:
    """Загружает HTML-страницу лида в /Stenvik/leads/ и публикует. Возвращает (remote_path, public_url)."""
    slug = _slugify(company_name)
    stamp = datetime.now().strftime("%y%m%d%H%M")
    remote_path = f"{LEADS_FOLDER}/{slug}-{stamp}.html"
    _upload_bytes(remote_path, html.encode("utf-8"))
    public_url = _publish_and_get_url(remote_path)
    logger.info("Uploaded + published HTML lead: %s → %s", remote_path, public_url)
    return remote_path, public_url


# ==== xlsx-дашборд ====

def _apply_sheet_style(ws, max_rows_for_validation: int = 10000) -> None:
    for col, title, width in COLUMNS:
        cell = ws[f"{col}1"]
        cell.value = title
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 28
    # Закрепить шапку + колонки Дата, Компания (свобода горизонтального скролла начинается с C)
    ws.freeze_panes = "C2"
    # Скрыть dedup_key
    ws.column_dimensions["L"].hidden = True

    dv_call = DataValidation(
        type="list", formula1=f'"{",".join(CALL_VALUES)}"',
        allow_blank=True, showDropDown=False,
    )
    dv_call.add(f"G2:G{max_rows_for_validation}")
    ws.add_data_validation(dv_call)

    dv_status = DataValidation(
        type="list", formula1=f'"{",".join(DEAL_STATUSES)}"',
        allow_blank=True, showDropDown=False,
    )
    dv_status.add(f"J2:J{max_rows_for_validation}")
    ws.add_data_validation(dv_status)

    last_col = COLUMNS[-1][0]
    ws.auto_filter.ref = f"A1:{last_col}1"


def download_sheet() -> Path:
    remote = settings.yandex_disk_file_path
    url = _get_download_url(remote)
    tmp = Path(tempfile.mkdtemp()) / "leads.xlsx"
    if url is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Лиды"
        _apply_sheet_style(ws)
        wb.save(tmp)
        return tmp
    with httpx.Client(timeout=60.0, follow_redirects=True) as client:
        r = client.get(url)
        r.raise_for_status()
        tmp.write_bytes(r.content)
    return tmp


def upload_sheet(local_path: Path) -> None:
    remote = settings.yandex_disk_file_path
    with open(local_path, "rb") as f:
        _upload_bytes(remote, f.read())


def append_lead_row(
    *,
    company_name: str,
    phone: str | None,
    city: str | None,
    priority: int,
    md_public_url: str,
    dedup_key: str,
) -> int:
    """Дописывает строку в таблицу. Детали лежат в MD-файле, здесь только краткое."""
    local = download_sheet()
    wb = load_workbook(local)
    ws = wb.active

    if ws.max_row == 0 or (ws.max_row == 1 and not any(c.value for c in ws[1])):
        _apply_sheet_style(ws)

    row = ws.max_row + 1
    values = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),  # A Дата
        company_name,                                 # B Компания
        phone or "",                                  # C Телефон
        city or "",                                   # D Город
        priority,                                     # E Приоритет
        "📄 Открыть",                                 # F Детали (hyperlink)
        "",                                           # G Позвонил
        "",                                           # H Дата звонка
        "",                                           # I Ответственный
        "Новый",                                      # J Статус сделки
        "",                                           # K Фидбэк
        dedup_key,                                    # L dedup_key
    ]
    for i, val in enumerate(values):
        col = get_column_letter(i + 1)
        cell = ws[f"{col}{row}"]
        cell.value = val
        cell.border = BORDER
        if col == "K":
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell.alignment = Alignment(vertical="center", horizontal="center" if col in ("E", "F", "G") else "left")
        if col == "F" and md_public_url:
            cell.hyperlink = md_public_url
            cell.font = LINK_FONT

    # Раскраска строки по приоритету
    fill = PRIORITY_FILLS.get(priority)
    if fill:
        for col, _, _ in COLUMNS:
            if col == "L":
                continue
            ws[f"{col}{row}"].fill = fill

    ws.row_dimensions[row].height = 22

    wb.save(local)
    upload_sheet(local)
    return row


def publish_sheet() -> str:
    """Возвращает публичный URL для xlsx."""
    return _publish_and_get_url(settings.yandex_disk_file_path)


# ==== мобильный HTML-дашборд ====

DASHBOARD_PATH = "/Stenvik/dashboard.html"


def _format_phone_tel(phone: str) -> str:
    """tel:-ссылка принимает только цифры (и +)."""
    cleaned = re.sub(r"[^\d+]", "", phone)
    return cleaned


def _priority_color(priority: int) -> tuple[str, str, str]:
    """Возвращает (bg_color, text_color, emoji)."""
    return {
        5: ("#fee2e2", "#991b1b", "🔴"),
        4: ("#fef3c7", "#92400e", "🟠"),
        3: ("#d1fae5", "#065f46", "🟢"),
        2: ("#f3f4f6", "#4b5563", "⚪"),
        1: ("#f3f4f6", "#4b5563", "⚪"),
    }.get(priority, ("#f3f4f6", "#4b5563", "⚪"))


def _esc(s) -> str:
    """HTML-escape."""
    if s is None:
        return ""
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


_WEBSITE_STATUS_LABEL = {
    "has_site": "Современный сайт",
    "cms_site": "Сайт на CMS (WordPress/Bitrix)",
    "constructor_site": "Сайт на конструкторе (Tilda/Wix)",
    "no_https": "Сайт без HTTPS",
    "dead_site": "Сайт не открывается",
}


def _shared_css() -> str:
    """Общий CSS для дашборда и страниц лидов — чтобы смотрелись единым сайтом."""
    return """
  :root {
    --bg: #f4f6fa;
    --surface: #ffffff;
    --text: #0f172a;
    --text-muted: #64748b;
    --border: #e2e8f0;
    --brand: #1f2937;
    --accent: #10b981;
    --link: #3b82f6;
    --p5-bg: #fee2e2; --p5-fg: #991b1b;
    --p4-bg: #fef3c7; --p4-fg: #92400e;
    --p3-bg: #d1fae5; --p3-fg: #065f46;
    --p2-bg: #f1f5f9; --p2-fg: #475569;
    --radius: 14px;
    --shadow-sm: 0 1px 3px rgba(15,23,42,0.06);
    --shadow-md: 0 4px 12px rgba(15,23,42,0.08);
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    background: var(--bg);
    color: var(--text);
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
    padding-bottom: 40px;
  }
  a { color: inherit; }
  .site-header {
    background: var(--brand);
    color: #fff;
    padding: 14px 16px 12px;
    position: sticky;
    top: 0;
    z-index: 20;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
  }
  .site-header-inner {
    max-width: 1280px;
    margin: 0 auto;
    display: flex;
    align-items: center;
    gap: 12px;
  }
  .site-header h1 {
    margin: 0;
    font-size: 17px;
    font-weight: 600;
    line-height: 1.2;
  }
  .site-header .upd {
    font-size: 12px;
    color: #9ca3af;
    margin-top: 3px;
  }
  .countdown {
    display: inline-block;
    background: rgba(255,255,255,0.12);
    border-radius: 6px;
    padding: 1px 7px;
    margin-left: 4px;
    font-variant-numeric: tabular-nums;
    letter-spacing: 0.02em;
  }
  .refreshing { animation: blink 0.4s step-end 3; }
  @keyframes blink { 50% { opacity: 0.35; } }
  .back-link {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    color: #e5e7eb;
    font-size: 14px;
    text-decoration: none;
    padding: 6px 10px 6px 6px;
    border-radius: 8px;
    transition: background 0.15s;
  }
  .back-link:hover, .back-link:active { background: rgba(255,255,255,0.1); }
  main { padding: 16px; max-width: 1280px; margin: 0 auto; }
  @media (min-width: 1024px) { main { padding: 24px; } }
"""


def _refresh_script() -> str:
    """JS, который перезагружает страницу с cache-busting каждые 3 мин."""
    # \\d → будет \d в JS regex (Python не ругается на unknown escape)
    return """
(function () {
  var INTERVAL = 3 * 60;
  var remaining = INTERVAL;
  var el = document.getElementById('cd');
  function fmt(s) {
    var m = Math.floor(s / 60);
    var sec = s % 60;
    return 'обновление через ' + m + ':' + (sec < 10 ? '0' : '') + sec;
  }
  function hardReload() {
    var url = location.href.replace(/[?&]_t=\\d+/, '');
    url += (url.indexOf('?') === -1 ? '?' : '&') + '_t=' + Date.now();
    location.replace(url);
  }
  var timer = setInterval(function () {
    remaining -= 1;
    if (remaining <= 0) {
      clearInterval(timer);
      if (el) { el.textContent = 'обновляем...'; el.classList.add('refreshing'); }
      setTimeout(hardReload, 600);
    } else if (el) {
      el.textContent = fmt(remaining);
    }
  }, 1000);
})();
"""


def build_lead_page_html(
    *,
    company_name: str,
    website_url: str | None,
    city: str | None,
    industry: str | None,
    website_status: str,
    phone: str | None,
    priority: int,
    priority_reason: str,
    summary: str,
    pains: list[str],
    recommended_services: list[str],
    sales_hook: str,
    analyzed_at: str = "",
    dashboard_url: str = "https://yadi.sk/d/3BppLUPjUVpQcg",
) -> str:
    """Рендерит полноценную HTML-страницу с деталями лида. Навигация назад — к дашборду."""
    bg, fg, emoji = _priority_color(priority)

    phone_raw = phone or ""
    phone_tel = _format_phone_tel(phone_raw) if phone_raw else ""

    actions_parts = []
    if phone_raw and phone_tel:
        actions_parts.append(
            f'<a class="btn btn-call" href="tel:{phone_tel}">'
            f'<span class="ic">📞</span><span>{_esc(phone_raw)}</span></a>'
        )
    if website_url:
        actions_parts.append(
            f'<a class="btn btn-site" href="{_esc(website_url)}" target="_blank" rel="noopener">'
            f'<span class="ic">🌐</span><span>Открыть сайт</span></a>'
        )
    actions_html = "".join(actions_parts) or '<div class="no-actions">Контактов нет — ищите компанию по названию</div>'

    pains_html = "".join(f'<li>{_esc(p)}</li>' for p in (pains or []))
    services_html = "".join(f'<span class="svc-chip">{_esc(s)}</span>' for s in (recommended_services or []))

    meta_parts = [x for x in [city, industry] if x]
    meta = " · ".join(_esc(x) for x in meta_parts)

    status_label = _WEBSITE_STATUS_LABEL.get(website_status, website_status or "")

    date_str = analyzed_at or datetime.now().strftime("%Y-%m-%d %H:%M")

    css = _shared_css() + f"""
  .lead-hero {{
    background: linear-gradient(135deg, {bg} 0%, #ffffff 100%);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 20px;
    margin-bottom: 16px;
  }}
  .lead-hero .prio-badge {{
    display: inline-block;
    background: {bg};
    color: {fg};
    font-weight: 700;
    font-size: 13px;
    padding: 4px 10px;
    border-radius: 999px;
    margin-bottom: 10px;
  }}
  .lead-hero h1 {{
    margin: 0 0 6px;
    font-size: 24px;
    font-weight: 700;
    line-height: 1.2;
    color: var(--text);
  }}
  @media (min-width: 640px) {{
    .lead-hero h1 {{ font-size: 28px; }}
    .lead-hero {{ padding: 24px; }}
  }}
  .lead-hero .meta {{ font-size: 14px; color: var(--text-muted); margin-bottom: 8px; }}
  .lead-hero .reason {{ font-size: 14px; color: var(--text); margin-top: 8px; font-style: italic; opacity: 0.85; }}
  .actions-row {{
    display: grid;
    grid-template-columns: 1fr;
    gap: 10px;
    margin-top: 16px;
  }}
  @media (min-width: 480px) {{ .actions-row {{ grid-template-columns: repeat(2, 1fr); }} }}
  .btn {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    padding: 14px;
    border-radius: 10px;
    font-size: 15px;
    font-weight: 600;
    text-decoration: none;
    min-height: 48px;
  }}
  .btn-call {{ background: var(--accent); color: #fff; }}
  .btn-call:active {{ background: #059669; }}
  .btn-site {{ background: var(--link); color: #fff; }}
  .btn-site:active {{ background: #2563eb; }}
  .no-actions {{
    background: #f8fafc; border: 1px dashed var(--border); border-radius: 10px;
    padding: 12px; text-align: center; color: var(--text-muted); font-size: 13px;
  }}
  .section {{
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 18px;
    margin-bottom: 14px;
    box-shadow: var(--shadow-sm);
  }}
  .section h2 {{
    margin: 0 0 12px;
    font-size: 15px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    color: var(--text-muted);
  }}
  .section p {{ margin: 0; font-size: 16px; line-height: 1.6; color: var(--text); }}
  .pains {{ margin: 0; padding-left: 22px; }}
  .pains li {{ margin-bottom: 8px; font-size: 15px; line-height: 1.5; }}
  .svc-grid {{ display: flex; flex-wrap: wrap; gap: 8px; }}
  .svc-chip {{
    background: #f1f5f9;
    color: var(--text);
    padding: 8px 14px;
    border-radius: 999px;
    font-size: 14px;
    font-weight: 500;
    border: 1px solid var(--border);
  }}
  .hook-box {{
    background: linear-gradient(135deg, #fef3c7 0%, #fff 100%);
    border: 1px solid #fde68a;
    border-radius: var(--radius);
    padding: 18px;
    margin-bottom: 14px;
  }}
  .hook-box h2 {{
    margin: 0 0 10px; font-size: 15px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.04em; color: #92400e;
  }}
  .hook-box blockquote {{
    margin: 0; font-size: 16px; line-height: 1.55; color: var(--text);
    border-left: 3px solid #f59e0b; padding-left: 14px;
  }}
  .meta-grid {{
    display: grid; grid-template-columns: 1fr; gap: 8px;
  }}
  @media (min-width: 640px) {{ .meta-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
  .meta-item {{ font-size: 13px; }}
  .meta-item .lbl {{ color: var(--text-muted); display: block; margin-bottom: 2px; }}
  .meta-item .val {{ color: var(--text); font-weight: 500; }}
  .page-footer {{
    text-align: center; color: var(--text-muted); font-size: 12px;
    padding: 20px 0 10px;
  }}
"""

    site_link = ""
    if website_url:
        site_link = f'<a href="{_esc(website_url)}" target="_blank" rel="noopener">{_esc(website_url)}</a>'

    reqs_parts = []
    if site_link:
        reqs_parts.append(f'<div class="meta-item"><span class="lbl">Сайт</span><span class="val">{site_link}</span></div>')
    if phone_raw:
        reqs_parts.append(f'<div class="meta-item"><span class="lbl">Телефон</span><span class="val">{_esc(phone_raw)}</span></div>')
    if city:
        reqs_parts.append(f'<div class="meta-item"><span class="lbl">Город</span><span class="val">{_esc(city)}</span></div>')
    if industry:
        reqs_parts.append(f'<div class="meta-item"><span class="lbl">Индустрия</span><span class="val">{_esc(industry)}</span></div>')
    if status_label:
        reqs_parts.append(f'<div class="meta-item"><span class="lbl">Статус сайта</span><span class="val">{_esc(status_label)}</span></div>')
    reqs_html = "".join(reqs_parts)

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{_esc(company_name)} — лид Stenvik</title>
<style>{css}</style>
</head>
<body>
<header class="site-header">
  <div class="site-header-inner">
    <a class="back-link" href="{_esc(dashboard_url)}">← К списку лидов</a>
  </div>
</header>
<main>
  <section class="lead-hero">
    <span class="prio-badge">{emoji} Приоритет {priority}/5</span>
    <h1>{_esc(company_name)}</h1>
    {f'<div class="meta">{meta}</div>' if meta else ''}
    {f'<div class="reason">{_esc(priority_reason)}</div>' if priority_reason else ''}
    <div class="actions-row">{actions_html}</div>
  </section>

  {f'<section class="section"><h2>О компании</h2><p>{_esc(summary)}</p></section>' if summary else ''}

  {f'<section class="section"><h2>Реквизиты</h2><div class="meta-grid">{reqs_html}</div></section>' if reqs_html else ''}

  {f'<section class="section"><h2>Боли — что не так с цифровизацией</h2><ol class="pains">{pains_html}</ol></section>' if pains_html else ''}

  {f'<section class="section"><h2>Что предложить (услуги Stenvik)</h2><div class="svc-grid">{services_html}</div></section>' if services_html else ''}

  {f'<section class="hook-box"><h2>Хук для продажника</h2><blockquote>{_esc(sales_hook)}</blockquote></section>' if sales_hook else ''}

  <div class="page-footer">Дата анализа: {_esc(date_str)} · Stenvik Lead Pipeline</div>
</main>
</body>
</html>
"""
    return html


def build_dashboard_html(leads: list[dict]) -> str:
    """Рендерит адаптивный HTML-дашборд: 1 колонка на мобиле, 2 на планшете, 3+ на десктопе.

    leads: список dict с полями: company_name, phone, city, industry, priority,
    priority_reason, summary, recommended_services, website_url, md_public_url,
    analyzed_at.
    """
    leads = sorted(leads, key=lambda x: x.get("analyzed_at") or "", reverse=True)
    leads = sorted(leads, key=lambda x: x.get("priority") or 0, reverse=True)

    total = len(leads)
    counts = {p: sum(1 for l in leads if l.get("priority") == p) for p in (5, 4, 3, 2, 1)}

    cards_html = []
    for lead in leads:
        p = int(lead.get("priority") or 0)
        bg, fg, emoji = _priority_color(p)
        phone_raw = lead.get("phone") or ""
        phone_tel = _format_phone_tel(phone_raw) if phone_raw else ""
        services = lead.get("recommended_services") or []
        services_html = "".join(
            f'<span class="svc">{_esc(s)}</span>' for s in services[:3]
        )

        md_url = lead.get("md_public_url") or ""
        site_url = lead.get("website_url") or ""

        company = _esc(lead.get("company_name") or "")
        summary = _esc(lead.get("summary") or "")
        city = lead.get("city") or ""
        industry = lead.get("industry") or ""
        meta_parts = []
        if city:
            meta_parts.append(_esc(city))
        if industry:
            meta_parts.append(_esc(industry))
        meta = " · ".join(meta_parts)

        # Кнопки на карточке
        buttons = []
        if phone_raw and phone_tel:
            buttons.append(
                f'<a class="btn-inline btn-call" href="tel:{phone_tel}" '
                f'onclick="event.stopPropagation()">📞 {_esc(phone_raw)}</a>'
            )
        if md_url:
            # details = главная ссылка, она же — «провал» в страницу лида
            buttons.append(
                f'<a class="btn-inline btn-details" href="{_esc(md_url)}" '
                f'onclick="event.stopPropagation()">Подробнее →</a>'
            )
        elif site_url:
            buttons.append(
                f'<a class="btn-inline btn-site" href="{_esc(site_url)}" target="_blank" rel="noopener" '
                f'onclick="event.stopPropagation()">🌐 Сайт</a>'
            )
        buttons_html = "".join(buttons)

        # Вся карточка — кликабельна: ведёт на страницу лида (md_url — это HTML-лендинг у новых,
        # MD-файл у старых; оба открываются одинаково в браузере)
        card_href = md_url or site_url
        card = f"""
          <article class="card" data-href="{_esc(card_href)}"{' onclick="cardClick(event, this)"' if card_href else ''}>
            <div class="card-top" style="background:{bg};">
              <div class="prio" style="color:{fg};">{emoji} {p}/5</div>
              <h2 class="name">{company}</h2>
              {f'<div class="meta">{meta}</div>' if meta else ''}
            </div>
            <div class="card-body">
              {f'<p class="summary">{summary}</p>' if summary else ''}
              {f'<div class="services">{services_html}</div>' if services else ''}
              {f'<div class="btn-row">{buttons_html}</div>' if buttons_html else ''}
            </div>
          </article>
        """
        cards_html.append(card)

    summary_inner = (
        f'<span class="pill p5">🔴 {counts.get(5, 0)}</span>'
        f'<span class="pill p4">🟠 {counts.get(4, 0)}</span>'
        f'<span class="pill p3">🟢 {counts.get(3, 0)}</span>'
        f'<span class="pill p2">⚪ {counts.get(2, 0) + counts.get(1, 0)}</span>'
        f'<span class="pill total">Всего: {total}</span>'
    )

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    shared = _shared_css()
    dashboard_css = """
  .summary-row {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    position: sticky;
    top: 54px;
    z-index: 15;
    padding: 10px 16px;
  }
  .summary-inner {
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    max-width: 1280px;
    margin: 0 auto;
  }
  .pill {
    font-size: 13px;
    padding: 5px 11px;
    border-radius: 999px;
    font-weight: 600;
  }
  .pill.p5 { background: var(--p5-bg); color: var(--p5-fg); }
  .pill.p4 { background: var(--p4-bg); color: var(--p4-fg); }
  .pill.p3 { background: var(--p3-bg); color: var(--p3-fg); }
  .pill.p2 { background: var(--p2-bg); color: var(--p2-fg); }
  .pill.total { background: var(--brand); color: #fff; }

  .grid {
    display: grid;
    grid-template-columns: 1fr;
    gap: 12px;
  }
  @media (min-width: 640px)  { .grid { grid-template-columns: repeat(2, 1fr); gap: 14px; } }
  @media (min-width: 1024px) { .grid { grid-template-columns: repeat(3, 1fr); gap: 16px; } }
  @media (min-width: 1440px) { .grid { grid-template-columns: repeat(4, 1fr); } }

  .card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    box-shadow: var(--shadow-sm);
    cursor: pointer;
    transition: transform 0.15s ease, box-shadow 0.15s ease, border-color 0.15s ease;
    display: flex;
    flex-direction: column;
  }
  .card:hover {
    transform: translateY(-2px);
    box-shadow: var(--shadow-md);
    border-color: #cbd5e1;
  }
  .card:active { transform: translateY(0); }
  .card-top { padding: 14px 16px 12px; }
  .prio {
    font-size: 12px;
    font-weight: 700;
    letter-spacing: 0.03em;
    margin-bottom: 4px;
  }
  .name {
    margin: 0 0 4px;
    font-size: 17px;
    font-weight: 700;
    color: var(--text);
    line-height: 1.25;
  }
  .meta {
    font-size: 13px;
    color: var(--text-muted);
  }
  .card-body { padding: 12px 16px 16px; display: flex; flex-direction: column; gap: 10px; flex: 1; }
  .summary {
    font-size: 14px;
    color: #334155;
    margin: 0;
    line-height: 1.5;
    display: -webkit-box;
    -webkit-line-clamp: 3;
    -webkit-box-orient: vertical;
    overflow: hidden;
  }
  .services {
    display: flex;
    flex-wrap: wrap;
    gap: 5px;
  }
  .svc {
    font-size: 11px;
    background: #f1f5f9;
    color: #475569;
    padding: 3px 8px;
    border-radius: 999px;
    border: 1px solid var(--border);
  }
  .btn-row {
    display: flex;
    gap: 8px;
    margin-top: auto;
    padding-top: 4px;
  }
  .btn-inline {
    flex: 1;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 6px;
    padding: 11px 12px;
    border-radius: 9px;
    font-size: 14px;
    font-weight: 600;
    text-decoration: none;
    min-height: 44px;
  }
  .btn-call { background: var(--accent); color: #fff; }
  .btn-call:active { background: #059669; }
  .btn-site { background: var(--link); color: #fff; }
  .btn-details { background: #eef2f7; color: var(--text); }
  .btn-details:active { background: #dfe5ee; }
  .empty {
    text-align: center;
    color: var(--text-muted);
    padding: 60px 20px;
    font-size: 15px;
  }
"""

    body_main = (
        f'<div class="grid">{"".join(cards_html)}</div>'
        if cards_html
        else '<div class="empty">Пока нет лидов — агент ищет первые компании.</div>'
    )

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Лиды Stenvik — дашборд продажника</title>
<style>{shared}{dashboard_css}</style>
</head>
<body>
<header class="site-header">
  <div class="site-header-inner" style="flex-direction:column; align-items:flex-start; gap:0;">
    <h1>Лиды Stenvik</h1>
    <div class="upd">Данные: {now} · <span id="cd" class="countdown">обновление через 3:00</span></div>
  </div>
</header>
<div class="summary-row"><div class="summary-inner">{summary_inner}</div></div>
<main>
{body_main}
</main>
<script>
function cardClick(e, el) {{
  if (e.target.closest('a')) return;
  var href = el.getAttribute('data-href');
  if (href) window.location.href = href;
}}
{_refresh_script()}
</script>
</body>
</html>
"""
    return html


def regenerate_dashboard(leads: list[dict]) -> str:
    """Генерирует HTML-дашборд, загружает на Я.Диск, публикует, возвращает public URL."""
    html = build_dashboard_html(leads)
    _upload_bytes(DASHBOARD_PATH, html.encode("utf-8"))
    return _publish_and_get_url(DASHBOARD_PATH)
